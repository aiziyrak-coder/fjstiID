"""
1-kurs.xlsx → backend/app/data/fjsti_students_1kurs.json
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

DEFAULT_XLSX = Path(r"c:\Users\alocomputers\Desktop\talablar kotingenti\talablar kotingenti\1-kurs.xlsx")
OUT = Path(__file__).resolve().parents[1] / "app" / "data" / "fjsti_students_1kurs.json"


def split_fio(full: str) -> tuple[str, str, str | None]:
    full = re.sub(r"\s+", " ", (full or "").strip())
    parts = full.split()
    if not parts:
        return "", "", None
    # oxiridagi qizi / o'g'li — otasining ismi qismi emas
    while parts and re.fullmatch(r"qizi|o['ʻ’`]?g['ʻ’`]?li|ogli", parts[-1], re.I):
        parts.pop()
    if not parts:
        return "", "", None
    if len(parts) == 1:
        return parts[0].title(), "-", None
    if len(parts) == 2:
        return parts[0].title(), parts[1].title(), None
    return parts[0].title(), parts[1].title(), " ".join(p.title() for p in parts[2:])


def gender_from_name(full: str) -> str | None:
    t = full.lower()
    if "qizi" in t or t.endswith("ovna") or "ovna" in t.split()[-1:]:
        return "female"
    if "o'g'li" in t or "ogli" in t.replace("'", "").replace("ʻ", "").replace("`", "") or t.endswith("ovich"):
        return "male"
    return None


def birth_gender_from_pinfl(pinfl: str) -> tuple[str | None, str | None]:
    """JSHSHIR dan tug'ilgan sana va jins (taxminiy)."""
    p = re.sub(r"\D", "", pinfl or "")
    if len(p) != 14:
        return None, None
    century = p[0]
    try:
        d, m, y = int(p[1:3]), int(p[3:5]), int(p[5:7])
    except ValueError:
        return None, None
    if century in ("1", "2"):
        year = 1800 + y
        gender = "male" if century == "1" else "female"
    elif century in ("3", "4"):
        year = 1900 + y
        gender = "male" if century == "3" else "female"
    elif century in ("5", "6"):
        year = 2000 + y
        gender = "male" if century == "5" else "female"
    else:
        return None, None
    if not (1 <= d <= 31 and 1 <= m <= 12 and 1930 <= year <= 2015):
        return None, gender
    return f"{year:04d}-{m:02d}-{d:02d}", gender


def split_passport(raw: str) -> tuple[str | None, str | None]:
    s = re.sub(r"\s+", "", (raw or "").upper())
    m = re.fullmatch(r"([A-Z]{2})(\d{7})", s)
    if m:
        return m.group(1), m.group(2)
    if s:
        return None, s
    return None, None


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not src.exists():
        raise SystemExit(f"Fayl topilmadi: {src}")

    wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
    ws = wb["Talabalar"] if "Talabalar" in wb.sheetnames else wb[wb.sheetnames[0]]

    people: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        full = str(row[0]).strip()
        passport = str(row[1]).strip() if row[1] is not None else ""
        student_id = str(row[2]).strip() if row[2] is not None else ""
        pinfl = str(row[3]).strip() if row[3] is not None else ""
        # openpyxl ba'zan int qaytaradi
        if pinfl.endswith(".0"):
            pinfl = pinfl[:-2]
        group = str(row[4]).strip() if row[4] is not None else ""

        last, first, middle = split_fio(full)
        if not last or not first:
            continue
        birth, g_pin = birth_gender_from_pinfl(pinfl)
        gender = gender_from_name(full) or g_pin
        series, number = split_passport(passport)

        people.append(
            {
                "full_name_raw": full,
                "last_name": last,
                "first_name": first,
                "middle_name": middle,
                "student_number": student_id or None,
                "pinfl": pinfl or None,
                "passport_series": series,
                "passport_number": number,
                "group_name": group or None,
                "birth_date": birth,
                "gender": gender,
                "course": 1,
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(people, ensure_ascii=False, indent=2), encoding="utf-8")
    groups = len({p["group_name"] for p in people if p["group_name"]})
    print(f"OK: {len(people)} talaba, {groups} guruh -> {OUT}")


if __name__ == "__main__":
    main()
