"""
Kurs Excel (1/2/...) → backend/app/data/fjsti_students_{N}kurs.json

Ustunlar sarlavhadan aniqlanadi (1-kurs va 2-kurs tartibi farq qilishi mumkin).
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

DATA_DIR = Path(__file__).resolve().parents[1] / "app" / "data"
DEFAULT_DIR = Path(r"c:\Users\alocomputers\Desktop\talablar kotingenti\talablar kotingenti")


def split_fio(full: str) -> tuple[str, str, str | None]:
    full = re.sub(r"\s+", " ", (full or "").strip())
    parts = full.split()
    if not parts:
        return "", "", None
    while parts and re.fullmatch(r"qizi|o['ʻ’`]?g['ʻ’`]?li|ogli", parts[-1], re.I):
        parts.pop()
    if not parts:
        return "", "", None
    if len(parts) == 1:
        return parts[0].title(), "-", None
    if len(parts) == 2:
        return parts[0].title(), parts[1].title(), None
    return parts[0].title(), parts[1].title(), " ".join(p.title() for p in parts[2:])


def gender_from_name(full: str) -> str | None:
    t = full.lower()
    if "qizi" in t or t.endswith("ovna"):
        return "female"
    t2 = t.replace("'", "").replace("ʻ", "").replace("`", "").replace("’", "").replace("‘", "")
    if "ogli" in t2 or t.endswith("ovich"):
        return "male"
    return None


def birth_gender_from_pinfl(pinfl: str) -> tuple[str | None, str | None]:
    p = re.sub(r"\D", "", pinfl or "")
    if len(p) != 14:
        return None, None
    century = p[0]
    try:
        d, m, y = int(p[1:3]), int(p[3:5]), int(p[5:7])
    except ValueError:
        return None, None
    if century in ("1", "2"):
        year, gender = 1800 + y, "male" if century == "1" else "female"
    elif century in ("3", "4"):
        year, gender = 1900 + y, "male" if century == "3" else "female"
    elif century in ("5", "6"):
        year, gender = 2000 + y, "male" if century == "5" else "female"
    else:
        return None, None
    if not (1 <= d <= 31 and 1 <= m <= 12 and 1930 <= year <= 2015):
        return None, gender
    return f"{year:04d}-{m:02d}-{d:02d}", gender


def split_passport(raw: str) -> tuple[str | None, str | None]:
    s = re.sub(r"\s+", "", (raw or "").upper())
    m = re.fullmatch(r"([A-Z]{2})(\d{7})", s)
    if m:
        return m.group(1), m.group(2)
    if s:
        return None, s
    return None, None


def cell_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s.replace(".0", "").isdigit():
        s = s[:-2]
    return s


def norm_header(h: str) -> str:
    t = (h or "").lower()
    t = t.replace("‘", "'").replace("’", "'").replace("`", "'")
    t = re.sub(r"\s+", " ", t).strip()
    return t


def map_headers(row: tuple) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, cell in enumerate(row):
        h = norm_header(cell_str(cell))
        if not h:
            continue
        if "to'liq" in h or "fio" in h or h == "f.i.sh" or "ismi" in h:
            idx["name"] = i
        elif "pasport" in h:
            idx["passport"] = i
        elif "jshshir" in h or "pinfl" in h:
            idx["pinfl"] = i
        elif "talaba id" in h or h == "talaba id":
            idx["student_id"] = i
        elif h == "guruh" or h.startswith("guruh"):
            idx["group"] = i
        elif h == "kurs" or h.startswith("kurs"):
            idx["course_label"] = i
    return idx


def course_from_label(label: str, fallback: int) -> int:
    m = re.search(r"(\d+)", label or "")
    if m:
        n = int(m.group(1))
        if 1 <= n <= 7:
            return n
    return fallback


def parse_file(src: Path, course_fallback: int) -> list[dict]:
    wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
    ws = wb["Talabalar"] if "Talabalar" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    cols = map_headers(header)
    if "name" not in cols:
        raise SystemExit(f"Ism ustuni topilmadi: {src.name} · {header}")

    # 1-kurs eski tartib fallback
    if "student_id" not in cols and "pinfl" in cols:
        # ba'zan almashtirilgan — qo'shimcha tekshiruv kerak emas, sarlavha bo'yicha
        pass
    if "student_id" not in cols:
        # eski: name, passport, student_id, pinfl, group
        cols.setdefault("passport", 1)
        cols.setdefault("student_id", 2)
        cols.setdefault("pinfl", 3)
        cols.setdefault("group", 4)

    people: list[dict] = []
    for row in rows:
        if not row:
            continue
        full = cell_str(row[cols["name"]] if cols["name"] < len(row) else None)
        if not full:
            continue
        passport = cell_str(row[cols["passport"]]) if "passport" in cols else ""
        student_id = cell_str(row[cols["student_id"]]) if "student_id" in cols else ""
        pinfl = cell_str(row[cols["pinfl"]]) if "pinfl" in cols else ""
        group = cell_str(row[cols["group"]]) if "group" in cols else ""
        course_label = cell_str(row[cols["course_label"]]) if "course_label" in cols else ""
        course = course_from_label(course_label, course_fallback)

        # agar student_id va pinfl almashtirilgan bo'lsa (14 raqam vs 12)
        if student_id and pinfl:
            sid_d = re.sub(r"\D", "", student_id)
            pin_d = re.sub(r"\D", "", pinfl)
            if len(sid_d) == 14 and len(pin_d) != 14 and len(pin_d) >= 10:
                student_id, pinfl = pinfl, student_id

        last, first, middle = split_fio(full)
        if not last or not first:
            continue
        birth, g_pin = birth_gender_from_pinfl(pinfl)
        gender = gender_from_name(full) or g_pin
        series, number = split_passport(passport)

        people.append(
            {
                "full_name_raw": full,
                "last_name": last,
                "first_name": first,
                "middle_name": middle,
                "student_number": student_id or None,
                "pinfl": re.sub(r"\D", "", pinfl)[:14] or None,
                "passport_series": series,
                "passport_number": number,
                "group_name": group or None,
                "birth_date": birth,
                "gender": gender,
                "course": course,
            }
        )
    return people


def main() -> None:
    # usage: parse_kurs.py [path.xlsx] [course_number]
    if len(sys.argv) >= 2:
        src = Path(sys.argv[1])
        course = int(sys.argv[2]) if len(sys.argv) >= 3 else course_from_label(src.stem, 1)
    else:
        src = DEFAULT_DIR / "2-kurs.xlsx"
        course = 2

    if not src.exists():
        raise SystemExit(f"Fayl topilmadi: {src}")

    people = parse_file(src, course)
    out = DATA_DIR / f"fjsti_students_{course}kurs.json"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(people, ensure_ascii=False, indent=2), encoding="utf-8")
    groups = len({p["group_name"] for p in people if p["group_name"]})
    print(f"OK: {len(people)} talaba, {groups} guruh, kurs={course} -> {out}")


if __name__ == "__main__":
    main()
